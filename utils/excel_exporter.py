"""
Excel Exporter - Export deal analysis to Excel
"""

import logging
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

def export_deals(deals, output_file):
    """Export deals to Excel with formatting and charts"""
    if not deals:
        logger.warning("No deals to export")
        return False
    
    try:
        # Convert deals to DataFrame
        df = pd.DataFrame([
            {
                'Score': deal['score'],
                'Address': deal['address'],
                'List Price': deal['list_price'],
                'ARV': deal['arv'],
                'Repair Costs': deal['repair_costs'],
                'Closing Costs': deal['closing_costs'],
                'Holding Costs': deal['holding_costs'],
                'Total Investment': deal['total_project_cost'],
                'Profit': deal['potential_profit'],
                'ROI (%)': deal['roi'],
                'Max Purchase Price (70% Rule)': deal['max_purchase_price'],
                'Meets 70% Rule': deal['meets_70_percent_rule'],
                'DOM': deal['property_data']['days_on_market'],
                'Bedrooms': deal['property_data']['bedrooms'],
                'Bathrooms': deal['property_data']['bathrooms'],
                'Square Feet': deal['property_data']['square_feet'],
                'Year Built': deal['property_data']['year_built'],
                'Price/SqFt': deal['property_data']['price_per_sqft']
            } for deal in deals
        ])
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # Save to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Deal Analysis', index=False)
            
            # Access the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Deal Analysis']
            
            # Set column widths
            for idx, col in enumerate(df.columns):
                column_width = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                worksheet.column_dimensions[get_column_letter(idx+1)].width = min(column_width, 30)
            
            # Add formatting
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_font = Font(bold=True)
            centered_alignment = Alignment(horizontal='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = centered_alignment
                cell.border = border
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = border
                    
                    # Center specific columns
                    if cell.column in [1, 11, 12, 13, 14, 15]:  # Score, Boolean, and count columns
                        cell.alignment = centered_alignment
                    
                    # Color ROI cells based on value
                    if cell.column == 10:  # ROI column
                        try:
                            roi_value = float(cell.value)
                            if roi_value >= 30:
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                            elif roi_value >= 20:
                                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
                            else:
                                cell.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Light pink
                        except (ValueError, TypeError):
                            pass
            
            # Add a bar chart for top properties by profit
            chart_sheet = workbook.create_sheet(title='Charts')
            
            profit_chart = BarChart()
            profit_chart.title = "Top Properties by Potential Profit"
            profit_chart.y_axis.title = "Profit ($)"
            profit_chart.x_axis.title = "Property"
            
            # Get top 10 properties by profit
            top_profits = df.sort_values('Profit', ascending=False).head(10)
            
            # Create references to data
            address_data = Reference(worksheet, min_col=2, min_row=2, max_row=min(11, len(df)+1))
            profit_data = Reference(worksheet, min_col=9, min_row=1, max_row=min(11, len(df)+1))
            
            profit_chart.add_data(profit_data, titles_from_data=True)
            profit_chart.set_categories(address_data)
            
            # Add chart to sheet
            chart_sheet.add_chart(profit_chart, "A1")
            
            # Add a second chart for ROI
            roi_chart = BarChart()
            roi_chart.title = "Top Properties by ROI"
            roi_chart.y_axis.title = "ROI (%)"
            roi_chart.x_axis.title = "Property"
            
            # Get top 10 properties by ROI
            top_roi = df.sort_values('ROI (%)', ascending=False).head(10)
            
            # Create references to data
            roi_data = Reference(worksheet, min_col=10, min_row=1, max_row=min(11, len(df)+1))
            
            roi_chart.add_data(roi_data, titles_from_data=True)
            roi_chart.set_categories(address_data)
            
            # Add chart to sheet
            chart_sheet.add_chart(roi_chart, "A20")
            
            # Add a summary sheet
            summary_sheet = workbook.create_sheet(title='Summary')
            
            # Summary statistics
            summary_data = {
                'Metric': [
                    'Total Properties Analyzed',
                    'Properties Meeting Criteria',
                    'Average List Price',
                    'Average ARV',
                    'Average Repair Costs',
                    'Average Profit',
                    'Average ROI',
                    'Highest Potential Profit',
                    'Highest ROI',
                    'Properties Meeting 70% Rule'
                ],
                'Value': [
                    len(df),
                    sum(1 for deal in deals if deal.get('meets_criteria', False)),
                    df['List Price'].mean(),
                    df['ARV'].mean(),
                    df['Repair Costs'].mean(),
                    df['Profit'].mean(),
                    df['ROI (%)'].mean(),
                    df['Profit'].max(),
                    df['ROI (%)'].max(),
                    sum(1 for deal in deals if deal.get('meets_70_percent_rule', False))
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(summary_sheet, index=False)
            
            # Format summary sheet
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = centered_alignment
                cell.border = border
            
            # Format the values
            for row in summary_sheet.iter_rows(min_row=2, max_row=len(summary_data['Metric'])+1, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border
                    
                    # Format numeric values
                    if cell.column == 2 and cell.row >= 3 and cell.row <= 9:  # Value column for monetary amounts
                        cell.number_format = '$#,##0.00'
                    elif cell.column == 2 and (cell.row == 10 or cell.row == 11):  # Percentage values
                        cell.number_format = '0.00%'
            
            # Adjust column widths in summary sheet
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
            
        # Log success
        logger.info(f"Exported {len(deals)} deals to Excel file: {output_file}")
        return True
    
    except Exception as e:
        logger.error(f"Error exporting deals to Excel: {str(e)}")
        return False
